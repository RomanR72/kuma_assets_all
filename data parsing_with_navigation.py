import pandas as pd
import json
import ast
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
import traceback
from tqdm import tqdm

def clean_tenant_name(tenant_name):
    """Очищает имя tenant от квадратных скобок и лишних пробелов"""
    if pd.isna(tenant_name):
        return None
    tenant_name = str(tenant_name)
    tenant_name = re.sub(r'\[.*?\]', '', tenant_name)
    tenant_name = ' '.join(tenant_name.split())
    return tenant_name.strip()

def add_back_to_navigation_link(ws):
    """Добавляет ссылку для возврата к навигационному листу"""
    back_cell = ws.cell(row=1, column=1, value="← Назад к навигации")
    back_cell.hyperlink = Hyperlink(ref="A1", location="'НАВИГАЦИЯ'!A1")
    back_cell.font = Font(color="0563C1", underline="single", bold=True)
    ws.row_dimensions[1].height = 20

def create_navigation_sheet(wb, tenants):
    """Создает навигационный лист с гиперссылками"""
    ws_nav = wb.create_sheet("НАВИГАЦИЯ", 0)
    
    # Заголовок
    ws_nav['A1'] = "НАВИГАЦИЯ ПО ОТЧЕТУ"
    ws_nav['A1'].font = Font(bold=True, size=14)
    
    # Заголовки столбцов
    headers = ["Tenant Name", "Main Data", "Software", "Vulnerabilities"]
    for col, header in enumerate(headers, 1):
        cell = ws_nav.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Данные
    for row_idx, tenant in enumerate(tenants, 3):
        clean_name = clean_tenant_name(tenant)
        prefix = clean_name[:25]
        
        ws_nav.cell(row=row_idx, column=1, value=clean_name)
        
        for col_idx, sheet_type in enumerate(["MAIN", "SOFTWARE", "VULNERABILITIES"], 2):
            sheet_name = f"{prefix}_{sheet_type}"
            if sheet_name in wb.sheetnames:
                cell = ws_nav.cell(row=row_idx, column=col_idx, value="Перейти")
                cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1")
                cell.font = Font(color="0563C1", underline="single")
    
    # Форматирование
    for column in ws_nav.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2 if any(cell.value for cell in column) else 0
        if max_length > 0:
            column_letter = column[0].column_letter
            ws_nav.column_dimensions[column_letter].width = max_length * 1.2
    
    # Чередующаяся заливка строк
    for row in range(3, len(tenants) + 3):
        if row % 2 == 1:
            for col in range(1, 5):
                ws_nav.cell(row=row, column=col).fill = PatternFill(
                    start_color="EFEFEF", end_color="EFEFEF", fill_type="solid"
                )

def parse_os_data(os_value):
    """Функция для парсинга данных ОС"""
    if pd.isna(os_value) or os_value == '':
        return None, None
    
    if isinstance(os_value, str):
        if not any(x in os_value.lower() for x in ['name', 'version']):
            parts = os_value.rsplit(' ', 1)
            return (parts[0], parts[1]) if len(parts) > 1 else (os_value, None)
        
        try:
            os_value = json.loads(os_value)
        except json.JSONDecodeError:
            try:
                os_value = ast.literal_eval(os_value)
            except:
                return str(os_value), None
    
    if isinstance(os_value, dict):
        return (
            str(os_value.get('name', '')).strip() or None,
            str(os_value.get('version', '')).strip() or None
        )
    return str(os_value), None

def parse_software_data(software_value):
    """Функция для парсинга данных о ПО"""
    if pd.isna(software_value) or software_value == '':
        return []
    
    if isinstance(software_value, str):
        try:
            software_value = json.loads(software_value)
        except json.JSONDecodeError:
            try:
                software_value = ast.literal_eval(software_value)
            except:
                return []
    
    if isinstance(software_value, list):
        return [{
            'name': str(item.get('name', '')).strip() or None,
            'version': str(item.get('version', '')).strip() or None,
            'vendor': str(item.get('vendor', '')).strip() or None
        } for item in software_value if isinstance(item, dict)]
    return []

def parse_vulnerabilities_data(vuln_value):
    """Функция для парсинга данных об уязвимостях"""
    if pd.isna(vuln_value) or vuln_value == '':
        return []
    
    if isinstance(vuln_value, str):
        try:
            vuln_value = json.loads(vuln_value)
        except json.JSONDecodeError:
            try:
                vuln_value = ast.literal_eval(vuln_value)
            except:
                return []
    
    if isinstance(vuln_value, list):
        return [{
            'kasperskyID': str(item.get('kasperskyID', '')).strip() or None,
            'productName': str(item.get('productName', '')).strip() or None,
            'descriptionURL': str(item.get('descriptionURL', '')).strip() or None,
            'recommendedMajorPatch': str(item.get('recommendedMajorPatch', '')).strip() or None,
            'recommendedMinorPatch': str(item.get('recommendedMinorPatch', '')).strip() or None,
            'severityStr': str(item.get('severityStr', '')).strip() or None,
            'severity': str(item.get('severity', '')).strip() or None,
            'cve': str(item.get('cve', '')).strip() or None,
            'exploitExists': str(item.get('exploitExists', '')).strip() or None,
            'malwareExists': str(item.get('malwareExists', '')).strip() or None
        } for item in vuln_value if isinstance(item, dict)]
    return []

def create_tenant_sheets(wb, merged_data, tenant_name):
    """Создает все листы для конкретного tenant"""
    try:
        clean_name = clean_tenant_name(tenant_name)
        prefix = clean_name[:25]
        tenant_data = merged_data[merged_data['tenantName'] == tenant_name].copy()
        
        # Лист MAIN
        main_columns = [
            'fqdn', 'ipAddresses', 'CPU', 'RAM', 'Disk Space',
            'Network Cards', 'macAddresses', 'os_parsed', 'os_version'
        ]
        main_data = tenant_data[[col for col in main_columns if col in tenant_data.columns]]
        main_data = main_data.rename(columns={'os_parsed': 'os'})
        
        ws_main = wb.create_sheet(f"{prefix}_MAIN")
        add_back_to_navigation_link(ws_main)
        for r in dataframe_to_rows(main_data, index=False, header=True):
            ws_main.append(r)
        
        # Лист SOFTWARE
        if 'software' in tenant_data.columns:
            software_rows = []
            for _, row in tenant_data[['fqdn', 'software']].iterrows():
                items = parse_software_data(row['software'])
                for item in items or [{}]:
                    software_rows.append({
                        'fqdn': row['fqdn'],
                        **item
                    })
            
            software_data = pd.DataFrame(software_rows)
            if not software_data.empty:
                ws_software = wb.create_sheet(f"{prefix}_SOFTWARE")
                add_back_to_navigation_link(ws_software)
                create_merged_sheet(ws_software, software_data, ['fqdn'], ['name', 'version', 'vendor'])
        
        # Лист VULNERABILITIES
        if 'vulnerabilities' in tenant_data.columns:
            vuln_rows = []
            for _, row in tenant_data[['fqdn', 'vulnerabilities']].iterrows():
                items = parse_vulnerabilities_data(row['vulnerabilities'])
                for item in items or [{}]:
                    vuln_rows.append({
                        'fqdn': row['fqdn'],
                        **item
                    })
            
            vuln_data = pd.DataFrame(vuln_rows)
            if not vuln_data.empty:
                ws_vuln = wb.create_sheet(f"{prefix}_VULNERABILITIES")
                add_back_to_navigation_link(ws_vuln)
                vuln_columns = [
                    'kasperskyID', 'productName', 'descriptionURL',
                    'recommendedMajorPatch', 'recommendedMinorPatch', 'severityStr',
                    'severity', 'cve', 'exploitExists', 'malwareExists'
                ]
                create_merged_sheet(ws_vuln, vuln_data, ['fqdn'], vuln_columns)
    
    except Exception as e:
        print(f"Ошибка создания листов для tenant {tenant_name}: {e}\n{traceback.format_exc()}")

def create_merged_sheet(ws, df, id_columns, data_columns):
    """Создает лист с объединенными ячейками для группированных данных"""
    try:
        # Пропускаем первую строку (там ссылка на навигацию)
        start_row = 2
        
        # Заголовки
        for col, header in enumerate(id_columns + data_columns, 1):
            ws.cell(row=start_row, column=col, value=header).font = Font(bold=True)
        
        if df.empty:
            return
        
        # Группировка и запись данных
        grouped = df.groupby(id_columns)
        current_row = start_row + 1
        
        for group_key, group in grouped:
            # Первая строка группы
            first_row = group.iloc[0]
            row_data = [first_row.get(col) for col in id_columns] + \
                      [first_row.get(col) for col in data_columns]
            ws.append(row_data)
            
            # Остальные строки группы
            for i in range(1, len(group)):
                row_data = [None]*len(id_columns) + \
                          [group.iloc[i].get(col) for col in data_columns]
                ws.append(row_data)
            
            # Объединение ячеек
            if len(group) > 1:
                for col in range(1, len(id_columns)+1):
                    ws.merge_cells(
                        start_row=current_row,
                        end_row=current_row + len(group) - 1,
                        start_column=col,
                        end_column=col
                    )
            
            current_row += len(group)
        
        # Центрирование объединенных ячеек
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=len(id_columns)):
            for cell in row:
                cell.alignment = Alignment(vertical='center', horizontal='center')
    
    except Exception as e:
        print(f"Ошибка при создании листа {ws.title}: {e}\n{traceback.format_exc()}")

def process_data(all_data_file, devices_report_file, output_file):
    try:
        print("Загрузка данных...")
        all_data = pd.read_excel(all_data_file)
        devices_report = pd.read_excel(devices_report_file)
        
        required_columns = ['fqdn', 'Network Cards', 'CPU', 'RAM', 'Disk Space']
        missing_columns = [col for col in required_columns if col not in devices_report.columns]
        if missing_columns:
            raise Exception(f"Отсутствуют столбцы: {', '.join(missing_columns)}")
        
        print("Объединение данных...")
        merged_data = pd.merge(
            all_data,
            devices_report[required_columns],
            on='fqdn',
            how='left'
        )
        
        if 'tenantName' in merged_data.columns:
            merged_data['tenantName'] = merged_data['tenantName'].apply(clean_tenant_name)
        
        if 'os' in merged_data.columns:
            print("Парсинг данных OS...")
            merged_data[['os_parsed', 'os_version']] = pd.DataFrame(
                merged_data['os'].apply(parse_os_data).tolist(),
                index=merged_data.index
            )
        
        if 'tenantName' not in merged_data.columns:
            raise Exception("Отсутствует столбец tenantName")
        
        print("Создание Excel файла...")
        wb = Workbook()
        wb.remove(wb.active)
        
        tenants = [t for t in merged_data['tenantName'].unique() if pd.notna(t)]
        
        print("Создание листов для каждого tenant...")
        for tenant in tqdm(tenants, desc="Обработка tenants"):
            create_tenant_sheets(wb, merged_data, tenant)
        
        print("Создание навигационного листа...")
        create_navigation_sheet(wb, tenants)
        
        print("Сохранение результата...")
        wb.save(output_file)
        print(f"\nФайл успешно сохранен: {output_file}")
        print(f"Создано листов: {len(wb.sheetnames)}")
        print(f"Обработано tenants: {len(tenants)}")
        
    except Exception as e:
        print(f"\nКритическая ошибка: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    try:
        input_files = {
            'all_data_combined': 'all_data_combined.xlsx',
            'devices_report': 'devices_report.xlsx'
        }
        
        missing_files = [name for name, path in input_files.items() if not pd.io.common.file_exists(path)]
        if missing_files:
            raise Exception(f"Отсутствуют файлы: {', '.join(missing_files)}")
        
        output_file = 'tenant_report_with_navigation.xlsx'
        
        print("Начало обработки данных...")
        process_data(input_files['all_data_combined'], 
                    input_files['devices_report'], 
                    output_file)
        
    except Exception as e:
        print(f"\nОшибка выполнения: {e}\n{traceback.format_exc()}")