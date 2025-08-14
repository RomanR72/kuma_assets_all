import pandas as pd
import json
import ast
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def parse_software_data(software_value):
    """Функция для парсинга данных о ПО"""
    if pd.isna(software_value):
        return []
    try:
        if isinstance(software_value, str):
            try:
                software_list = json.loads(software_value)
            except json.JSONDecodeError:
                try:
                    software_list = ast.literal_eval(software_value)
                except:
                    return []
        if isinstance(software_list, list):
            return [{
                'software': str(item.get('name', '')).strip(),
                'version': str(item.get('version', '')).strip(),
                'vendor': str(item.get('vendor', '')).strip()
            } for item in software_list if any(item.values())]
        return []
    except Exception as e:
        print(f"Ошибка парсинга ПО: {e}")
        return []

def parse_vulnerabilities_data(vuln_value):
    """Функция для парсинга данных об уязвимостях"""
    if pd.isna(vuln_value):
        return []
    try:
        if isinstance(vuln_value, str):
            try:
                vuln_list = json.loads(vuln_value)
            except json.JSONDecodeError:
                try:
                    vuln_list = ast.literal_eval(vuln_value)
                except:
                    return []
        if isinstance(vuln_list, list):
            return [{
                'kasperskyID': str(item.get('kasperskyID', '')).strip(),
                'productName': str(item.get('productName', '')).strip(),
                'descriptionURL': str(item.get('descriptionURL', '')).strip(),
                'recommendedMajorPatch': str(item.get('recommendedMajorPatch', '')).strip(),
                'recommendedMinorPatch': str(item.get('recommendedMinorPatch', '')).strip(),
                'severityStr': str(item.get('severityStr', '')).strip(),
                'severity': str(item.get('severity', '')).strip(),
                'cve': str(item.get('cve', '')).strip(),
                'exploitExists': str(item.get('exploitExists', '')).strip(),
                'malwareExists': str(item.get('malwareExists', '')).strip()
            } for item in vuln_list if any(item.values())]
        return []
    except Exception as e:
        print(f"Ошибка парсинга уязвимостей: {e}")
        return []

def parse_os_data(os_value):
    """Функция для парсинга данных ОС"""
    if pd.isna(os_value):
        return None, None
    try:
        if isinstance(os_value, str) and not any(x in os_value for x in ['name', 'version']):
            parts = os_value.rsplit(' ', 1)
            return (parts[0], parts[1]) if len(parts) > 1 else (os_value, None)
        
        if isinstance(os_value, str):
            try:
                os_dict = json.loads(os_value)
            except json.JSONDecodeError:
                os_dict = ast.literal_eval(os_value)
        else:
            os_dict = os_value
        
        return (
            str(os_dict.get('name', '')).strip() or None,
            str(os_dict.get('version', '')).strip() or None
        )
    except Exception as e:
        print(f"Ошибка парсинга OS: {e}")
        return str(os_value), None

def process_data(all_data_file, devices_report_file, output_file):
    # Загрузка и объединение данных
    all_data = pd.read_excel(all_data_file)
    devices_report = pd.read_excel(devices_report_file)[['fqdn', 'Network Cards', 'CPU', 'RAM', 'Disk Space']]
    merged_data = pd.merge(all_data, devices_report, on='fqdn', how='left')

    # Парсинг OS
    if 'os' in merged_data.columns:
        merged_data[['os_parsed', 'os_version']] = pd.DataFrame(
            merged_data['os'].apply(parse_os_data).tolist(), index=merged_data.index)

    # Подготовка основного листа
    main_columns = [
        'tenantName', 'fqdn', 'ipAddresses', 'CPU', 'RAM', 'Disk Space',
        'Network Cards', 'macAddresses', 'os_parsed', 'os_version'
    ]
    main_sheet = merged_data[main_columns].rename(columns={'os_parsed': 'os'})

    # Подготовка листа software
    software_data = merged_data[['tenantName', 'fqdn', 'software']]
    software_rows = []
    for _, row in software_data.iterrows():
        items = parse_software_data(row['software'])
        for item in items or [{}]:
            software_rows.append({
                'tenantName': row['tenantName'],
                'fqdn': row['fqdn'],
                **{k: v for k, v in item.items() if v}
            })
    software_sheet = pd.DataFrame(software_rows)

    # Подготовка листа vulnerabilities
    vuln_data = merged_data[['tenantName', 'fqdn', 'vulnerabilities']]
    vuln_rows = []
    for _, row in vuln_data.iterrows():
        items = parse_vulnerabilities_data(row['vulnerabilities'])
        for item in items or [{}]:
            vuln_rows.append({
                'tenantName': row['tenantName'],
                'fqdn': row['fqdn'],
                **{k: v for k, v in item.items() if v}
            })
    vuln_sheet = pd.DataFrame(vuln_rows)

    # Создание Excel файла
    wb = Workbook()
    wb.remove(wb.active)

    # Лист main
    ws_main = wb.create_sheet("main")
    for r in dataframe_to_rows(main_sheet, index=False, header=True):
        ws_main.append(r)

    # Лист software с объединением ячеек
    ws_software = wb.create_sheet("software")
    ws_software.append(['tenantName', 'fqdn', 'software', 'version', 'vendor'])
    
    current_row = 2
    for fqdn, group in software_sheet.groupby('fqdn'):
        first_row = True
        for _, row in group.iterrows():
            if first_row:
                ws_software.append([
                    row['tenantName'],
                    row['fqdn'],
                    row.get('software'),
                    row.get('version'),
                    row.get('vendor')
                ])
                first_row = False
            else:
                ws_software.append([
                    None, None,
                    row.get('software'),
                    row.get('version'),
                    row.get('vendor')
                ])
        
        if len(group) > 1:
            ws_software.merge_cells(
                start_row=current_row,
                end_row=current_row + len(group) - 1,
                start_column=1,
                end_column=1
            )
            ws_software.merge_cells(
                start_row=current_row,
                end_row=current_row + len(group) - 1,
                start_column=2,
                end_column=2
            )
        current_row += len(group)

    # Лист vulnerabilities с объединением ячеек
    ws_vuln = wb.create_sheet("vulnerabilities")
    vuln_headers = [
        'tenantName', 'fqdn', 'kasperskyID', 'productName', 'descriptionURL',
        'recommendedMajorPatch', 'recommendedMinorPatch', 'severityStr',
        'severity', 'cve', 'exploitExists', 'malwareExists'
    ]
    ws_vuln.append(vuln_headers)
    
    current_row = 2
    for fqdn, group in vuln_sheet.groupby('fqdn'):
        first_row = True
        for _, row in group.iterrows():
            if first_row:
                ws_vuln.append([row.get(col) for col in vuln_headers])
                first_row = False
            else:
                ws_vuln.append([
                    None, None,
                    row.get('kasperskyID'),
                    row.get('productName'),
                    row.get('descriptionURL'),
                    row.get('recommendedMajorPatch'),
                    row.get('recommendedMinorPatch'),
                    row.get('severityStr'),
                    row.get('severity'),
                    row.get('cve'),
                    row.get('exploitExists'),
                    row.get('malwareExists')
                ])
        
        if len(group) > 1:
            ws_vuln.merge_cells(
                start_row=current_row,
                end_row=current_row + len(group) - 1,
                start_column=1,
                end_column=1
            )
            ws_vuln.merge_cells(
                start_row=current_row,
                end_row=current_row + len(group) - 1,
                start_column=2,
                end_column=2
            )
        current_row += len(group)

    # Центрирование объединенных ячеек
    for sheet in [ws_software, ws_vuln]:
        for row in sheet.iter_rows():
            for cell in row[:2]:  # Только для столбцов tenantName и fqdn
                cell.alignment = Alignment(vertical='center', horizontal='center')

    wb.save(output_file)
    print(f"Файл успешно сохранен: {output_file}")

# Запуск обработки
process_data('all_data_combined.xlsx', 'devices_report.xlsx', 'all_data_combined_final.xlsx')