import pandas as pd
import json
import ast
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
import traceback
from tqdm import tqdm

def clean_tenant_name(tenant_name):
    """Очищает имя tenant от квадратных скобок и лишних пробелов"""
    if pd.isna(tenant_name):
        return None
    if not isinstance(tenant_name, str):
        tenant_name = str(tenant_name)
    return re.sub(r'\[.*?\]', '', tenant_name).strip()

def add_back_to_navigation_link(ws):
    """Добавляет ссылку для возврата к навигационному листу"""
    ws.cell(row=1, column=1, value="← Назад к навигации").hyperlink = Hyperlink(
        ref="A1", location="'НАВИГАЦИЯ'!A1")
    ws.cell(row=1, column=1).font = Font(color="0563C1", underline="single", bold=True)
    ws.row_dimensions[1].height = 20

def create_navigation_sheet(wb, tenants):
    """Создает навигационный лист с гиперссылками"""
    ws_nav = wb.create_sheet("НАВИГАЦИЯ", 0)
    
    # Заголовок
    ws_nav['A1'] = "НАВИГАЦИЯ ПО ОТЧЕТУ"
    ws_nav['A1'].font = Font(bold=True, size=14)
    
    # Заголовки столбцов
    headers = ["Tenant Name", "Main Data", "Software", "Vulnerabilities"]
    for col, header in enumerate(headers, 1):
        cell = ws_nav.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Данные
    for row_idx, tenant in enumerate(tenants, 3):
        clean_name = clean_tenant_name(tenant)
        prefix = clean_name[:25]
        
        ws_nav.cell(row=row_idx, column=1, value=clean_name)
        
        for col_idx, sheet_type in enumerate(["MAIN", "SOFTWARE", "VULNERABILITIES"], 2):
            sheet_name = f"{prefix}_{sheet_type}"
            if sheet_name in wb.sheetnames:
                cell = ws_nav.cell(row=row_idx, column=col_idx, value="Перейти")
                cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1")
                cell.font = Font(color="0563C1", underline="single")
    
    # Оптимизированное форматирование
    for col in ws_nav.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
        if max_len > 0:
            ws_nav.column_dimensions[col[0].column_letter].width = (max_len + 2) * 1.2
    
    # Чередующаяся заливка строк
    for row in range(3, len(tenants) + 3):
        if row % 2 == 1:
            for col in range(1, 5):
                ws_nav.cell(row=row, column=col).fill = PatternFill(
                    start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

def parse_os_data(os_value):
    """Оптимизированный парсинг данных ОС"""
    if pd.isna(os_value) or not os_value:
        return None, None
    
    if isinstance(os_value, str):
        if 'name' not in os_value.lower() and 'version' not in os_value.lower():
            parts = os_value.rsplit(' ', 1)
            return (parts[0], parts[1]) if len(parts) > 1 else (os_value, None)
        
        try:
            os_dict = json.loads(os_value) if os_value.startswith(('{', '[')) else ast.literal_eval(os_value)
            if isinstance(os_dict, dict):
                return (
                    str(os_dict.get('name', '')).strip() or None,
                    str(os_dict.get('version', '')).strip() or None
                )
        except:
            return str(os_value), None
    
    if isinstance(os_value, dict):
        return (
            str(os_value.get('name', '')).strip() or None,
            str(os_value.get('version', '')).strip() or None
        )
    return str(os_value), None

def parse_json_data(value, fields):
    """Универсальная функция для парсинга JSON данных"""
    if pd.isna(value) or not value:
        return []
    
    try:
        if isinstance(value, str):
            try:
                data = json.loads(value)
            except json.JSONDecodeError:
                try:
                    data = ast.literal_eval(value)
                except:
                    return []
        else:
            data = value
        
        if isinstance(data, list):
            return [
                {field: str(item.get(field, '')).strip() or None for field in fields}
                for item in data if isinstance(item, dict)
            ]
    except:
        return []
    return []

def create_tenant_sheets(wb, merged_data, tenant_name):
    """Создает все листы для конкретного tenant"""
    try:
        clean_name = clean_tenant_name(tenant_name)
        prefix = clean_name[:25]
        tenant_data = merged_data[merged_data['tenantName'] == tenant_name].copy()
        
        # Лист MAIN
        main_columns = [
            'fqdn', 'ipAddresses', 'CPU', 'RAM', 'Disk Space',
            'Network Cards', 'macAddresses', 'os_parsed', 'os_version'
        ]
        main_data = tenant_data[[col for col in main_columns if col in tenant_data.columns]]
        main_data = main_data.rename(columns={'os_parsed': 'os'})
        
        ws_main = wb.create_sheet(f"{prefix}_MAIN")
        add_back_to_navigation_link(ws_main)
        for r in dataframe_to_rows(main_data, index=False, header=True):
            ws_main.append(r)
        
        # Лист SOFTWARE
        if 'software' in tenant_data.columns:
            software_rows = []
            for _, row in tenant_data[['fqdn', 'software']].iterrows():
                for item in parse_json_data(row['software'], ['name', 'version', 'vendor']):
                    software_rows.append({
                        'fqdn': row['fqdn'],
                        **item
                    })
            
            if software_rows:
                ws_software = wb.create_sheet(f"{prefix}_SOFTWARE")
                add_back_to_navigation_link(ws_software)
                create_merged_sheet(ws_software, pd.DataFrame(software_rows), 
                                  ['fqdn'], ['name', 'version', 'vendor'])
        
        # Лист VULNERABILITIES
        if 'vulnerabilities' in tenant_data.columns:
            vuln_fields = ['kasperskyID', 'productName', 'descriptionURL', 'recommendedMajorPatch',
                         'recommendedMinorPatch', 'severityStr', 'severity', 'cve', 'exploitExists', 'malwareExists']
            vuln_rows = []
            for _, row in tenant_data[['fqdn', 'vulnerabilities']].iterrows():
                for item in parse_json_data(row['vulnerabilities'], vuln_fields):
                    vuln_rows.append({
                        'fqdn': row['fqdn'],
                        **item
                    })
            
            if vuln_rows:
                ws_vuln = wb.create_sheet(f"{prefix}_VULNERABILITIES")
                add_back_to_navigation_link(ws_vuln)
                create_merged_sheet(ws_vuln, pd.DataFrame(vuln_rows), 
                                   ['fqdn'], vuln_fields)
    
    except Exception as e:
        print(f"Ошибка создания листов для tenant {tenant_name}: {e}\n{traceback.format_exc()}")

def create_merged_sheet(ws, df, id_columns, data_columns):
    """Создает лист с объединенными ячейками для группированных данных"""
    try:
        # Заголовки
        ws.append(id_columns + data_columns)
        
        if df.empty:
            return
        
        # Группировка и запись данных
        current_row = 2
        for _, group in df.groupby(id_columns):
            # Первая строка группы
            first_row = group.iloc[0]
            ws.append([first_row[col] for col in id_columns] + 
                     [first_row[col] for col in data_columns])
            
            # Остальные строки группы
            for i in range(1, len(group)):
                ws.append([None]*len(id_columns) + 
                         [group.iloc[i][col] for col in data_columns])
            
            # Объединение ячеек
            if len(group) > 1:
                for col in range(1, len(id_columns)+1):
                    ws.merge_cells(
                        start_row=current_row,
                        end_row=current_row + len(group) - 1,
                        start_column=col,
                        end_column=col
                    )
            
            current_row += len(group)
        
        # Центрирование объединенных ячеек
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(id_columns)):
            for cell in row:
                cell.alignment = Alignment(vertical='center', horizontal='center')
    
    except Exception as e:
        print(f"Ошибка при создании листа {ws.title}: {e}\n{traceback.format_exc()}")

def process_data(all_data_file, devices_report_file, output_file):
    try:
        print("Оптимизированная загрузка данных...")
        # Указываем только нужные столбцы и их типы для экономии памяти
        dtype = {
            'tenantName': 'category',
            'fqdn': 'category',
            'os': 'object',
            'software': 'object',
            'vulnerabilities': 'object'
        }
        all_data = pd.read_excel(all_data_file, dtype=dtype)
        devices_report = pd.read_excel(devices_report_file, 
                                     usecols=['fqdn', 'Network Cards', 'CPU', 'RAM', 'Disk Space'],
                                     dtype={'fqdn': 'category'})
        
        print("Объединение данных...")
        merged_data = pd.merge(
            all_data,
            devices_report,
            on='fqdn',
            how='left'
        )
        
        # Очистка tenantName
        if 'tenantName' in merged_data.columns:
            merged_data['tenantName'] = merged_data['tenantName'].apply(clean_tenant_name)
        
        # Парсинг OS
        if 'os' in merged_data.columns:
            print("Парсинг данных OS...")
            merged_data[['os_parsed', 'os_version']] = pd.DataFrame(
                merged_data['os'].apply(parse_os_data).tolist(),
                index=merged_data.index
            )
        
        if 'tenantName' not in merged_data.columns:
            raise Exception("Отсутствует столбец tenantName в исходных данных")
        
        print("Создание Excel файла...")
        wb = Workbook()
        wb.remove(wb.active)
        
        tenants = [t for t in merged_data['tenantName'].unique() if pd.notna(t)]
        
        print("Создание листов для каждого tenant...")
        for tenant in tqdm(tenants, desc="Обработка tenants"):
            create_tenant_sheets(wb, merged_data, tenant)
        
        print("Создание навигационного листа...")
        create_navigation_sheet(wb, tenants)
        
        print("Сохранение результата...")
        wb.save(output_file)
        print(f"\nФайл успешно сохранен: {output_file}")
        print(f"Создано листов: {len(wb.sheetnames)}")
        print(f"Обработано tenants: {len(tenants)}")
        
    except Exception as e:
        print(f"\nКритическая ошибка: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    try:
        input_files = {
            'all_data_combined': 'all_data_combined.xlsx',
            'devices_report': 'devices_report.xlsx'
        }
        
        missing_files = [name for name, path in input_files.items() if not pd.io.common.file_exists(path)]
        if missing_files:
            raise Exception(f"Отсутствуют файлы: {', '.join(missing_files)}")
        
        output_file = 'tenant_report_final.xlsx'
        
        print("Начало обработки данных...")
        process_data(input_files['all_data_combined'], 
                    input_files['devices_report'], 
                    output_file)
        
    except Exception as e:
        print(f"\nОшибка выполнения: {e}\n{traceback.format_exc()}")